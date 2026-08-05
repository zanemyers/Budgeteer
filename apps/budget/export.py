"""
Excel export for a budget.

The point is portability: someone picking this up, or walking away from it, should be able to open
one file and see the whole picture without the app. That pulls in two directions. A crosstab with
categories as columns is the readable shape — it is how a spreadsheet budget is normally kept — but
`Date, Description, Amount, Category` is what every other budgeting app's importer expects. One
sheet cannot be both, so both are written, from the same TransactionLine rows, and so cannot
disagree with each other.

Amounts are written as `amount_usd * user_rate`, which is what the dashboard displays, so the file
reconciles against the screen rather than against the raw per-transaction currency.
"""

import calendar
import datetime
from decimal import Decimal

from django.db.models import Sum
from django.db.models.functions import Coalesce

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.budget.data import get_budget_overview
from apps.budget.models import Category, TransactionLine

MONEY = "#,##0.00"
HEADER_FILL = PatternFill("solid", fgColor="E8EDE3")
SECTION_FILL = PatternFill("solid", fgColor="D6DFCC")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _autosize(sheet, minimum=10, maximum=42):
    """Width each column to its widest cell, clamped so one long note cannot swallow the sheet."""
    for column in sheet.columns:
        longest = max((len(str(cell.value)) for cell in column if cell.value is not None), default=0)
        letter = get_column_letter(column[0].column)
        sheet.column_dimensions[letter].width = max(minimum, min(maximum, longest + 2))


def _header_row(sheet, row, values):
    for index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=index, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="bottom", wrap_text=True)
    return row + 1


def _month_bounds(year: int, month: int) -> tuple[datetime.date, datetime.date]:
    last = calendar.monthrange(year, month)[1]
    return datetime.date(year, month, 1), datetime.date(year, month, last)


def _lines(budget, start: datetime.date, end: datetime.date):
    """
    Every transaction line whose effective date falls in the window.

    Effective date is paid_date falling back to due_date, matching how the transactions page and
    the dashboard decide which month a row belongs to, so the export agrees with both.
    """
    return (
        TransactionLine.objects.filter(transaction__budget=budget)
        .annotate(effective=Coalesce("transaction__paid_date", "transaction__due_date"))
        .filter(effective__range=(start, end))
        .select_related("transaction", "category", "transaction__payment_method")
        .order_by("effective", "transaction_id", "pk")
    )


def _amount(line, rate: Decimal) -> Decimal:
    return (line.amount_usd or Decimal("0")) * rate


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------


def _section(sheet, row, title, width):
    """
    Write a banded section heading.

    Filled across `width` columns rather than merged: a merged cell is the first thing to break
    when a sheet is saved as CSV or pulled into another tool, and this file is meant to travel.
    """
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=11)
    for column in range(1, width + 1):
        sheet.cell(row=row, column=column).fill = SECTION_FILL
    return row + 2


def _write_month_overview(sheet, budget, month_str: str, rate: Decimal, currency: str) -> None:
    """
    Write the month at a glance, with income and expenses as separate sections.

    They used to share one table distinguished only by a Type column, which read as a single
    undifferentiated block — and three of that table's columns are meaningless for income, since an
    income category has no assigned amount, no target and nothing remaining. Each section now shows
    only the columns that mean something for it.
    """
    overview = get_budget_overview(budget, month_str, rate)
    start, _ = _month_bounds(*(int(part) for part in month_str.split("-")))
    income = [c for c in overview["categories"] if c["category_type"] == Category.TYPE_INCOME]
    expenses = [c for c in overview["categories"] if c["category_type"] != Category.TYPE_INCOME and not c["is_goal"]]

    sheet["A1"] = f"{budget.name or f'Budget #{budget.pk}'} — {start:%B %Y}"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = f"All amounts in {currency}."

    # The summary sits at the top but reads from the section totals below it, so editing a row keeps
    # the whole sheet consistent. Its value cells are filled in once those rows are known.
    summary_start = _section(sheet, 4, "SUMMARY", 6)
    summary_labels = [
        "Income",
        "Expenses",
        "Saved to goals",
        "Spent from goals",
        "Assigned to categories",
        "Left to assign",
    ]
    for offset, label in enumerate(summary_labels):
        sheet.cell(row=summary_start + offset, column=1, value=label).font = HEADER_FONT
    row = summary_start + len(summary_labels) + 2

    row = _section(sheet, row, "INCOME", 2)
    row = _header_row(sheet, row, ["Category", "Received"])
    income_first = row
    for category in income:
        sheet.cell(row=row, column=1, value=category["name"])
        cell = sheet.cell(row=row, column=2, value=Decimal(category["activity"]))
        cell.number_format = MONEY
        row += 1
    income_total_row = row
    sheet.cell(row=row, column=1, value="TOTAL INCOME").font = HEADER_FONT
    cell = sheet.cell(row=row, column=2, value=f"=SUM(B{income_first}:B{row - 1})" if income else Decimal("0"))
    cell.number_format = MONEY
    cell.font = HEADER_FONT
    row += 3

    row = _section(sheet, row, "EXPENSES", 5)
    row = _header_row(sheet, row, ["Category", "Assigned", "Target", "Spent", "Remaining"])
    expense_first = row
    for category in expenses:
        sheet.cell(row=row, column=1, value=category["name"])
        for column, key in ((2, "assigned"), (3, "budgeted"), (4, "activity"), (5, "available")):
            cell = sheet.cell(row=row, column=column, value=Decimal(category[key]))
            cell.number_format = MONEY
        row += 1
    expense_total_row = row
    sheet.cell(row=row, column=1, value="TOTAL EXPENSES").font = HEADER_FONT
    for column in (2, 3, 4, 5):
        letter = get_column_letter(column)
        cell = sheet.cell(
            row=row,
            column=column,
            value=f"=SUM({letter}{expense_first}:{letter}{row - 1})" if expenses else Decimal("0"),
        )
        cell.number_format = MONEY
        cell.font = HEADER_FONT

    summary_values = [
        f"=B{income_total_row}",
        f"=D{expense_total_row}",
        Decimal(overview["saved_to_goals_total"]),
        Decimal(overview["goal_monthly_spending"]),
        f"=B{expense_total_row}",
        Decimal(overview["ready_to_assign"]),
    ]
    for offset, value in enumerate(summary_values):
        cell = sheet.cell(row=summary_start + offset, column=2, value=value)
        cell.number_format = MONEY

    _autosize(sheet)


def _write_year_overview(sheet, budget, year: int, rate: Decimal, currency: str) -> None:
    sheet["A1"] = f"{budget.name or f'Budget #{budget.pk}'} — {year}"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = f"All amounts in {currency}."

    row = _header_row(
        sheet,
        4,
        ["Month", "Income", "Expenses", "Assigned", "Saved to goals", "Spent from goals", "Left to assign"],
    )
    first_data_row = row
    for month in range(1, 13):
        overview = get_budget_overview(budget, f"{year}-{month:02d}", rate)
        expenses = sum((Decimal(c["activity"]) for c in overview["categories"] if not c["is_goal"]), Decimal())
        sheet.cell(row=row, column=1, value=f"{datetime.date(year, month, 1):%B}")
        values = (
            Decimal(overview["income_total"]),
            expenses,
            Decimal(overview["expense_assigned"]),
            Decimal(overview["saved_to_goals_total"]),
            Decimal(overview["goal_monthly_spending"]),
            Decimal(overview["ready_to_assign"]),
        )
        for offset, value in enumerate(values, start=2):
            cell = sheet.cell(row=row, column=offset, value=value)
            cell.number_format = MONEY
        row += 1

    sheet.cell(row=row, column=1, value="TOTAL").font = HEADER_FONT
    for column in range(2, 8):
        letter = get_column_letter(column)
        cell = sheet.cell(row=row, column=column, value=f"=SUM({letter}{first_data_row}:{letter}{row - 1})")
        cell.number_format = MONEY
        cell.font = HEADER_FONT

    sheet.freeze_panes = "A5"
    _autosize(sheet)


def _write_crosstab(sheet, lines, rate: Decimal, currency: str) -> None:
    """
    One row per transaction, one column per category it touched.

    This is the shape a hand-kept spreadsheet budget uses. What such a sheet normally hides in a
    cell comment — what the amount was for and when — is written as real columns here instead:
    a comment cannot be sorted, filtered or pivoted, and does not survive a save as CSV.
    """
    grouped: dict[int, dict] = {}
    for line in lines:
        txn = line.transaction
        entry = grouped.setdefault(
            txn.pk,
            {
                "date": line.effective,
                "description": txn.description,
                "type": txn.transaction_type or "",
                "method": str(txn.payment_method) if txn.payment_method else "",
                "notes": txn.notes,
                "amounts": {},
                "total": Decimal("0"),
            },
        )
        if line.category_id is not None:
            entry["amounts"][line.category_id] = entry["amounts"].get(line.category_id, Decimal("0")) + _amount(
                line, rate
            )
        entry["total"] += _amount(line, rate)

    used_ids = {cat_id for entry in grouped.values() for cat_id in entry["amounts"]}
    categories = list(Category.objects.filter(pk__in=used_ids).order_by("category_type", "name").values("pk", "name"))

    sheet["A1"] = f"All amounts in {currency}. A split transaction is one row with amounts under each category."
    fixed = ["Date", "Description", "Type", "Payment method", "Notes"]
    row = _header_row(sheet, 3, [*fixed, *(c["name"] for c in categories), "Row total"])
    first_data_row = row
    column_of = {c["pk"]: len(fixed) + index + 1 for index, c in enumerate(categories)}
    total_column = len(fixed) + len(categories) + 1

    for entry in grouped.values():
        sheet.cell(row=row, column=1, value=entry["date"]).number_format = "yyyy-mm-dd"
        sheet.cell(row=row, column=2, value=entry["description"])
        sheet.cell(row=row, column=3, value=entry["type"])
        sheet.cell(row=row, column=4, value=entry["method"])
        sheet.cell(row=row, column=5, value=entry["notes"])
        for cat_id, amount in entry["amounts"].items():
            cell = sheet.cell(row=row, column=column_of[cat_id], value=amount)
            cell.number_format = MONEY
        cell = sheet.cell(row=row, column=total_column, value=entry["total"])
        cell.number_format = MONEY
        row += 1

    if row > first_data_row:
        sheet.cell(row=row, column=1, value="TOTAL").font = HEADER_FONT
        for column in list(column_of.values()) + [total_column]:
            letter = get_column_letter(column)
            cell = sheet.cell(row=row, column=column, value=f"=SUM({letter}{first_data_row}:{letter}{row - 1})")
            cell.number_format = MONEY
            cell.font = HEADER_FONT

    sheet.freeze_panes = "C4"
    _autosize(sheet, maximum=30)


def _write_flat(sheet, lines, rate: Decimal, currency: str) -> None:
    """One row per line, in the column order other budgeting apps expect on import."""
    sheet["A1"] = f"All amounts in {currency}. One row per transaction line, for importing elsewhere."
    row = _header_row(
        sheet, 3, ["Date", "Description", "Category", "Category type", "Amount", "Payment method", "Notes", "Status"]
    )
    for line in lines:
        txn = line.transaction
        sheet.cell(row=row, column=1, value=line.effective).number_format = "yyyy-mm-dd"
        sheet.cell(row=row, column=2, value=line.description or txn.description)
        sheet.cell(row=row, column=3, value=line.category.name if line.category else "")
        sheet.cell(row=row, column=4, value=line.category.category_type if line.category else "")
        cell = sheet.cell(row=row, column=5, value=_amount(line, rate))
        cell.number_format = MONEY
        sheet.cell(row=row, column=6, value=str(txn.payment_method) if txn.payment_method else "")
        sheet.cell(row=row, column=7, value=txn.notes)
        sheet.cell(row=row, column=8, value="paid" if txn.paid_date else "pending")
        row += 1

    sheet.freeze_panes = "A4"
    _autosize(sheet, maximum=30)


def _write_goals(sheet, budget, month_str: str, rate: Decimal, currency: str) -> None:
    overview = get_budget_overview(budget, month_str, rate)
    goals = [c for c in overview["categories"] if c["is_goal"]]

    sheet["A1"] = f"Goals. All amounts in {currency}."
    row = _header_row(sheet, 3, ["Goal", "Target", "Saved", "Still to go", "Due", "Ongoing", "Months left"])
    for goal in goals:
        target = Decimal(goal["goal_target"] or "0")
        saved = Decimal(goal["goal_total_saved"] or "0")
        sheet.cell(row=row, column=1, value=goal["name"])
        for column, value in ((2, target), (3, saved), (4, target - saved)):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.number_format = MONEY
        sheet.cell(row=row, column=5, value=goal["goal_due_date"] or "")
        sheet.cell(row=row, column=6, value="yes" if goal["goal_ongoing"] else "no")
        sheet.cell(row=row, column=7, value=goal["goal_months_remaining"])
        row += 1

    if not goals:
        sheet.cell(row=row, column=1, value="No goals in this budget.")

    sheet.freeze_panes = "A4"
    _autosize(sheet)


def _write_goal_spending(sheet, budget, start: datetime.date, end: datetime.date, rate: Decimal, currency: str) -> None:
    """Expenses drawn against a goal balance, which come out of savings rather than the month."""
    lines = (
        _lines(budget, start, end)
        .filter(category__goal__isnull=False, transaction__transaction_type="expense")
        .order_by("effective", "pk")
    )

    sheet["A1"] = f"Money paid out of a goal. All amounts in {currency}."
    sheet["A2"] = "These draw on a goal's saved balance, so they are not funded by the month's income."
    row = _header_row(sheet, 4, ["Date", "Goal", "Description", "Amount", "Payment method", "Notes"])
    first_data_row = row
    for line in lines:
        txn = line.transaction
        sheet.cell(row=row, column=1, value=line.effective).number_format = "yyyy-mm-dd"
        sheet.cell(row=row, column=2, value=line.category.name if line.category else "")
        sheet.cell(row=row, column=3, value=line.description or txn.description)
        cell = sheet.cell(row=row, column=4, value=_amount(line, rate))
        cell.number_format = MONEY
        sheet.cell(row=row, column=5, value=str(txn.payment_method) if txn.payment_method else "")
        sheet.cell(row=row, column=6, value=txn.notes)
        row += 1

    if row > first_data_row:
        sheet.cell(row=row, column=1, value="TOTAL").font = HEADER_FONT
        cell = sheet.cell(row=row, column=4, value=f"=SUM(D{first_data_row}:D{row - 1})")
        cell.number_format = MONEY
        cell.font = HEADER_FONT
    else:
        sheet.cell(row=row, column=1, value="Nothing was paid out of a goal in this period.")

    sheet.freeze_panes = "A5"
    _autosize(sheet)


# ---------------------------------------------------------------------------
# Workbooks
# ---------------------------------------------------------------------------


def _totals_note(workbook) -> None:
    """Column totals are written as formulas so the file stays live if someone edits a row."""
    workbook.properties.creator = "Budgeteer"


def build_flat_workbook(transactions, rate: Decimal, currency: str) -> Workbook:
    """
    One sheet, one row per transaction line, for the rows currently on the transactions page.

    Takes an already-filtered Transaction queryset rather than a date window, so what downloads is
    exactly what was on screen — including a search, which spans every month.
    """
    lines = (
        TransactionLine.objects.filter(transaction__in=transactions)
        .annotate(effective=Coalesce("transaction__paid_date", "transaction__due_date"))
        .select_related("transaction", "category", "transaction__payment_method")
        .order_by("effective", "transaction_id", "pk")
    )
    workbook = Workbook()
    workbook.active.title = "Transactions"
    _write_flat(workbook.active, lines, rate, currency)
    _totals_note(workbook)
    return workbook


def build_month_workbook(budget, month_str: str, rate: Decimal, currency: str) -> Workbook:
    year, month = (int(part) for part in month_str.split("-"))
    start, end = _month_bounds(year, month)
    lines = list(_lines(budget, start, end))

    workbook = Workbook()
    _write_month_overview(workbook.active, budget, month_str, rate, currency)
    workbook.active.title = "Overview"
    _write_crosstab(workbook.create_sheet("Transactions"), lines, rate, currency)
    # No flat sheet here: the transactions page exports that shape for whatever is on screen, so
    # carrying a second copy of the same rows in this file only made it bigger. The year export
    # keeps one, because a full year of flat rows is not reachable from that page.
    _write_goals(workbook.create_sheet("Goals"), budget, month_str, rate, currency)
    _write_goal_spending(workbook.create_sheet("Paid from goals"), budget, start, end, rate, currency)
    _totals_note(workbook)
    return workbook


def build_year_workbook(budget, year: int, rate: Decimal, currency: str) -> Workbook:
    start = datetime.date(year, 1, 1)
    end = datetime.date(year, 12, 31)
    lines = list(_lines(budget, start, end))

    workbook = Workbook()
    _write_year_overview(workbook.active, budget, year, rate, currency)
    workbook.active.title = "Overview"
    _write_crosstab(workbook.create_sheet("Transactions"), lines, rate, currency)
    _write_flat(workbook.create_sheet("Transactions flat"), lines, rate, currency)
    # A goal's saved balance is cumulative, so December is the state at the end of the year.
    _write_goals(workbook.create_sheet("Goals"), budget, f"{year}-12", rate, currency)
    _write_goal_spending(workbook.create_sheet("Paid from goals"), budget, start, end, rate, currency)
    _totals_note(workbook)
    return workbook


def workbook_filename(budget, label: str) -> str:
    name = "".join(ch if ch.isalnum() or ch in " -_" else "" for ch in (budget.name or f"Budget {budget.pk}")).strip()
    return f"{name or 'Budget'} {label}.xlsx"


def aggregate_sum(budget, start, end, rate) -> Decimal:
    """Total of every line in the window, used by the tests to check the sheets tie out."""
    total = _lines(budget, start, end).aggregate(total=Sum("amount_usd"))["total"] or Decimal("0")
    return total * rate
