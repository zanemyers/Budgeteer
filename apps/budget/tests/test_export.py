"""
Tests for the spreadsheet export.

The export exists so the data is not trapped in the app, which makes "the numbers in the file are
the numbers in the app" the property worth pinning. The crosstab and the flat sheet are two views
of the same TransactionLine rows, so the strongest check is that they total to each other and to
the overview.
"""

import datetime
import io
from decimal import Decimal

from django.urls import reverse

from openpyxl import load_workbook

from apps.accounts.models import User
from apps.base.tests import BaseTest
from apps.budget.export import build_month_workbook, build_year_workbook, workbook_filename
from apps.budget.models import (
    Budget,
    BudgetMembership,
    Category,
    CategoryBudget,
    Goal,
    Transaction,
    TransactionLine,
)

# The month export carries no flat sheet: the transactions page exports that shape for whatever is
# on screen. The year export keeps one, because a full year of flat rows cannot be reached from that
# page — its month filter only widens for a search.
MONTH_SHEETS = ["Overview", "Transactions", "Goals", "Paid from goals"]
YEAR_SHEETS = ["Overview", "Transactions", "Transactions flat", "Goals", "Paid from goals"]


class ExportTestCase(BaseTest):
    def setUp(self):
        super().setUp()
        self.user = self.make_user()
        self.budget = Budget.objects.create(created_by=self.user, name="Household")
        BudgetMembership.objects.create(budget=self.budget, user=self.user, role=BudgetMembership.ROLE_OWNER)
        self.rent = Category.objects.create(budget=self.budget, name="Rent", category_type=Category.TYPE_EXPENSE)
        self.food = Category.objects.create(budget=self.budget, name="Groceries", category_type=Category.TYPE_EXPENSE)
        self.pay = Category.objects.create(budget=self.budget, name="Salary", category_type=Category.TYPE_INCOME)
        self.roof = Category.objects.create(budget=self.budget, name="New roof", category_type=Category.TYPE_EXPENSE)
        Goal.objects.create(category=self.roof, target=Decimal("9000.00"), due_date=datetime.date(2027, 1, 1))
        CategoryBudget.objects.create(
            budget=self.budget, category=self.rent, month=datetime.date(2026, 8, 1), assigned=Decimal("1500.00")
        )

        self._txn("Rent", datetime.date(2026, 8, 1), [(self.rent, "1500.00")], "expense")
        # A split: one transaction, two categories. The crosstab keeps it as a single row.
        self._txn("Costco", datetime.date(2026, 8, 6), [(self.food, "41.20"), (self.rent, "18.75")], "expense")
        self._txn("August pay", datetime.date(2026, 8, 15), [(self.pay, "2000.00")], "income")
        self._txn("Roof deposit", datetime.date(2026, 8, 20), [(self.roof, "250.00")], "expense")
        # Outside the month, so it must not appear in a month export.
        self._txn("July rent", datetime.date(2026, 7, 1), [(self.rent, "1500.00")], "expense")

    def _txn(self, description, date, lines, txn_type):
        txn = Transaction.objects.create(
            budget=self.budget,
            created_by=self.user,
            description=description,
            due_date=date,
            paid_date=date,
            transaction_type=txn_type,
        )
        for category, amount in lines:
            TransactionLine.objects.create(
                transaction=txn, category=category, amount=Decimal(amount), amount_usd=Decimal(amount)
            )
        return txn

    def _month(self, month="2026-08"):
        return build_month_workbook(self.budget, month, Decimal("1"), "USD")


class TestWorkbookShape(ExportTestCase):
    def test_the_month_sheets_are_present_and_in_order(self):
        self.assertEqual(self._month().sheetnames, MONTH_SHEETS)

    def test_the_month_export_carries_no_flat_sheet(self):
        """It is reachable from the transactions page, so a second copy only grew the file."""
        self.assertNotIn("Transactions flat", self._month().sheetnames)

    def test_the_year_export_keeps_a_flat_sheet(self):
        """A full year of flat rows is not obtainable from the transactions page."""
        self.assertEqual(build_year_workbook(self.budget, 2026, Decimal("1"), "USD").sheetnames, YEAR_SHEETS)

    def test_the_filename_carries_the_budget_name_and_period(self):
        self.assertEqual(workbook_filename(self.budget, "August 2026"), "Household August 2026.xlsx")

    def test_a_budget_name_with_punctuation_stays_a_usable_filename(self):
        self.budget.name = 'Zane/Ashlynn "joint": 2026'
        self.assertEqual(workbook_filename(self.budget, "2026"), "ZaneAshlynn joint 2026 2026.xlsx")

    def test_the_workbook_survives_a_round_trip_through_a_file(self):
        buffer = io.BytesIO()
        self._month().save(buffer)
        buffer.seek(0)
        self.assertEqual(load_workbook(buffer).sheetnames, MONTH_SHEETS)


class TestCrosstab(ExportTestCase):
    def _sheet(self):
        return self._month()["Transactions"]

    def _headers(self):
        sheet = self._sheet()
        return [cell.value for cell in sheet[3] if cell.value is not None]

    def test_categories_used_in_the_month_become_columns(self):
        headers = self._headers()
        for name in ("Rent", "Groceries", "Salary"):
            self.assertIn(name, headers)

    def test_a_category_with_no_activity_gets_no_column(self):
        unused = Category.objects.create(budget=self.budget, name="Unused", category_type=Category.TYPE_EXPENSE)
        self.assertNotIn(unused.name, self._headers())

    def test_the_metadata_is_in_real_columns_not_cell_comments(self):
        """A comment cannot be sorted, filtered or pivoted, and does not survive a save as CSV."""
        sheet = self._sheet()
        self.assertEqual(
            [cell.value for cell in sheet[3][:5]], ["Date", "Description", "Type", "Payment method", "Notes"]
        )
        self.assertTrue(all(cell.comment is None for row in sheet.iter_rows() for cell in row))

    def test_a_split_transaction_is_one_row_with_two_amounts(self):
        sheet = self._sheet()
        headers = [cell.value for cell in sheet[3]]
        rows = [r for r in sheet.iter_rows(min_row=4, values_only=True) if r[1] == "Costco"]
        self.assertEqual(len(rows), 1, "the split became two rows")
        row = rows[0]
        self.assertEqual(row[headers.index("Groceries")], Decimal("41.20"))
        self.assertEqual(row[headers.index("Rent")], Decimal("18.75"))

    def test_a_transaction_outside_the_month_is_absent(self):
        descriptions = [r[1] for r in self._sheet().iter_rows(min_row=4, values_only=True)]
        self.assertNotIn("July rent", descriptions)


class TestSheetsAgree(ExportTestCase):
    """
    The crosstab and the flat sheet are the same rows twice, so they have to total the same.

    Checked on the year workbook, which is the one carrying both since the month export dropped its
    flat sheet.
    """

    def _year(self):
        return build_year_workbook(self.budget, 2026, Decimal("1"), "USD")

    def _date_rows(self, sheet):
        return [r for r in sheet.iter_rows(min_row=4, values_only=True) if isinstance(r[0], datetime.date)]

    def test_the_two_views_of_the_same_rows_total_the_same(self):
        workbook = self._year()
        crosstab = workbook["Transactions"]
        total_column = [cell.value for cell in crosstab[3]].index("Row total")
        crosstab_total = sum((row[total_column] for row in self._date_rows(crosstab)), Decimal("0"))
        flat_total = sum((row[4] for row in self._date_rows(workbook["Transactions flat"])), Decimal("0"))

        self.assertEqual(crosstab_total, flat_total)
        # 1500 + 41.20 + 18.75 + 2000 + 250 in August, plus 1500 in July.
        self.assertEqual(flat_total, Decimal("5309.95"))

    def test_the_flat_sheet_has_one_row_per_line(self):
        rows = self._date_rows(self._year()["Transactions flat"])
        self.assertEqual(len(rows), 6, "five transactions, one of which is a two-line split")

    def test_the_month_crosstab_totals_only_that_month(self):
        crosstab = self._month()["Transactions"]
        total_column = [cell.value for cell in crosstab[3]].index("Row total")
        total = sum((row[total_column] for row in self._date_rows(crosstab)), Decimal("0"))
        self.assertEqual(total, Decimal("3809.95"), "July's rent leaked into an August export")


class TestGoalSheets(ExportTestCase):
    def test_goals_show_target_saved_and_what_is_left(self):
        sheet = self._month()["Goals"]
        self.assertEqual([cell.value for cell in sheet[3][:4]], ["Goal", "Target", "Saved", "Still to go"])
        row = next(r for r in sheet.iter_rows(min_row=4, values_only=True) if r[0] == "New roof")
        self.assertEqual(row[1], Decimal("9000.00"))
        self.assertEqual(row[1] - row[2], row[3], "still-to-go must be target minus saved")

    def test_money_paid_out_of_a_goal_is_listed(self):
        sheet = self._month()["Paid from goals"]
        rows = [r for r in sheet.iter_rows(min_row=5, values_only=True) if isinstance(r[0], datetime.date)]
        self.assertEqual([r[1] for r in rows], ["New roof"])
        self.assertEqual([r[3] for r in rows], [Decimal("250.00")])

    def test_a_budget_with_no_goals_says_so_rather_than_showing_an_empty_grid(self):
        Goal.objects.all().delete()
        sheet = self._month()["Goals"]
        self.assertEqual(sheet.cell(row=4, column=1).value, "No goals in this budget.")


class TestExportView(ExportTestCase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.user)
        self.url = reverse("budget:export", kwargs={"budget_pk": self.budget.pk})

    def test_a_month_download_is_an_xlsx_attachment(self):
        res = self.client.get(self.url + "?month=2026-08")
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn('attachment; filename="Household August 2026.xlsx"', res["Content-Disposition"])
        self.assertEqual(load_workbook(io.BytesIO(res.content)).sheetnames, MONTH_SHEETS)

    def test_a_year_download_names_the_year(self):
        res = self.client.get(self.url + "?year=2026")
        self.assertEqual(res.status_code, 200)
        self.assertIn('filename="Household 2026.xlsx"', res["Content-Disposition"])

    def test_a_bad_month_is_rejected_not_a_500(self):
        res = self.client.get(self.url + "?month=nonsense")
        self.assertEqual(res.status_code, 400)
        self.assertIn("month", res.json()["errors"])

    def test_a_bad_year_is_rejected_not_a_500(self):
        for value in ("abc", "12", "99999"):
            with self.subTest(year=value):
                res = self.client.get(self.url + f"?year={value}")
                self.assertEqual(res.status_code, 400)
                self.assertIn("year", res.json()["errors"])

    def test_someone_elses_budget_is_not_exportable(self):
        stranger = User.objects.create_user(username="other", email="other@example.com", password="pw")  # noqa: S106
        other_budget = Budget.objects.create(created_by=stranger)
        BudgetMembership.objects.create(budget=other_budget, user=stranger, role=BudgetMembership.ROLE_OWNER)
        url = reverse("budget:export", kwargs={"budget_pk": other_budget.pk})
        self.assertEqual(self.client.get(url + "?month=2026-08").status_code, 404)


class TestOverviewSheet(ExportTestCase):
    """
    Income and expenses have to read as separate sections.

    They shared one table distinguished only by a Type column, which read as a single
    undifferentiated block — and three of that table's columns are meaningless for income, since an
    income category has no assigned amount, no target and nothing remaining.
    """

    def _sheet(self):
        return self._month()["Overview"]

    def _find(self, sheet, label):
        for row in sheet.iter_rows(min_col=1, max_col=1):
            if row[0].value == label:
                return row[0].row
        raise AssertionError(f"{label!r} is not on the sheet")

    def test_income_and_expenses_are_separate_sections(self):
        sheet = self._sheet()
        income_row = self._find(sheet, "INCOME")
        expense_row = self._find(sheet, "EXPENSES")
        self.assertLess(income_row, expense_row, "income should come first, as the money arrives first")

    def test_the_income_section_only_shows_columns_that_mean_something(self):
        sheet = self._sheet()
        header = self._find(sheet, "Category")
        self.assertEqual([sheet.cell(row=header, column=c).value for c in (1, 2, 3)], ["Category", "Received", None])

    def test_the_expense_section_shows_the_envelope_columns(self):
        sheet = self._sheet()
        header = self._find(sheet, "EXPENSES") + 2
        self.assertEqual(
            [sheet.cell(row=header, column=c).value for c in range(1, 6)],
            ["Category", "Assigned", "Target", "Spent", "Remaining"],
        )

    def test_each_section_totals_itself(self):
        sheet = self._sheet()
        income_total = self._find(sheet, "TOTAL INCOME")
        expense_total = self._find(sheet, "TOTAL EXPENSES")
        self.assertTrue(str(sheet.cell(row=income_total, column=2).value).startswith("=SUM("))
        self.assertTrue(str(sheet.cell(row=expense_total, column=4).value).startswith("=SUM("))

    def test_the_summary_reads_from_the_sections_below_it(self):
        """So editing a category row keeps the top of the sheet honest instead of going stale."""
        sheet = self._sheet()
        income_total = self._find(sheet, "TOTAL INCOME")
        expense_total = self._find(sheet, "TOTAL EXPENSES")
        self.assertEqual(sheet.cell(row=self._find(sheet, "Income"), column=2).value, f"=B{income_total}")
        self.assertEqual(sheet.cell(row=self._find(sheet, "Expenses"), column=2).value, f"=D{expense_total}")

    def test_nothing_in_the_workbook_is_merged(self):
        """A merged cell is the first thing to break on a save as CSV, and this file travels."""
        workbook = self._month()
        for name in workbook.sheetnames:
            with self.subTest(sheet=name):
                self.assertEqual(list(workbook[name].merged_cells.ranges), [])


class TestTransactionExportView(ExportTestCase):
    """
    The transactions export has to be the rows on screen.

    It hangs off a page with a month, a category, a payment method, a date range and a search
    applied, so exporting anything other than what those filters produced would be a surprise. The
    filter predicate is shared with the page rather than duplicated, and these pin that.
    """

    def setUp(self):
        super().setUp()
        self.client.force_login(self.user)
        self.url = reverse("budget:transaction-export", kwargs={"budget_pk": self.budget.pk})

    def _rows(self, query):
        res = self.client.get(self.url + query)
        self.assertEqual(res.status_code, 200)
        sheet = load_workbook(io.BytesIO(res.content))["Transactions"]
        return [r for r in sheet.iter_rows(min_row=4, values_only=True) if isinstance(r[0], datetime.date)]

    def test_it_is_a_single_flat_sheet(self):
        res = self.client.get(self.url + "?month=2026-08")
        self.assertEqual(load_workbook(io.BytesIO(res.content)).sheetnames, ["Transactions"])

    def test_it_exports_the_month_on_screen(self):
        descriptions = [r[1] for r in self._rows("?month=2026-08")]
        self.assertNotIn("July rent", descriptions, "a July row escaped an August export")
        self.assertIn("Rent", descriptions)

    def test_a_category_filter_narrows_the_file(self):
        """
        Only transactions touching the category, but all of their lines.

        Costco is a Groceries/Rent split. The page filters transactions and shows each matching one
        whole, so its Rent line appears there too — and the file has to be the same rows as the page.
        Trimming the file to Groceries lines alone would make the two disagree.
        """
        rows = self._rows(f"?month=2026-08&category={self.food.pk}")
        self.assertEqual({r[1] for r in rows}, {"Costco"}, "only the transaction touching Groceries")
        self.assertEqual(sorted(r[2] for r in rows), ["Groceries", "Rent"], "both sides of the split")

    def test_a_search_widens_the_file_past_the_month(self):
        """Search spans every month on the page, so it has to here too."""
        descriptions = [r[1] for r in self._rows("?month=2026-08&q=rent")]
        self.assertIn("July rent", descriptions)

    def test_a_date_range_narrows_the_file(self):
        rows = self._rows("?month=2026-08&date_from=2026-08-10&date_to=2026-08-31")
        self.assertEqual(sorted({r[1] for r in rows}), ["August pay", "Roof deposit"])

    def test_the_filename_names_a_search_rather_than_a_month(self):
        res = self.client.get(self.url + "?month=2026-08&q=rent")
        self.assertIn("transactions matching rent", res["Content-Disposition"])

    def test_a_bad_month_is_rejected_not_a_500(self):
        res = self.client.get(self.url + "?month=nonsense")
        self.assertEqual(res.status_code, 400)
        self.assertIn("month", res.json()["errors"])

    def test_someone_elses_budget_is_not_exportable(self):
        stranger = User.objects.create_user(username="other2", email="other2@example.com", password="pw")  # noqa: S106
        other = Budget.objects.create(created_by=stranger)
        BudgetMembership.objects.create(budget=other, user=stranger, role=BudgetMembership.ROLE_OWNER)
        url = reverse("budget:transaction-export", kwargs={"budget_pk": other.pk})
        self.assertEqual(self.client.get(url + "?month=2026-08").status_code, 404)

    def test_the_export_and_the_page_return_the_same_rows(self):
        """The one property worth guarding: the file and the screen come from one predicate."""
        query = f"?month=2026-08&category={self.rent.pk}"
        page = self.client.get(
            reverse("budget:transaction-list", kwargs={"budget_pk": self.budget.pk}) + query,
            headers={"x-inertia": "true", "x-inertia-version": "1.0"},
        )
        on_screen = {t["description"] for t in page.json()["props"]["transactions"]}
        in_file = {r[1] for r in self._rows(query)}
        self.assertEqual(in_file, on_screen)
